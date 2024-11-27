import aiohttp
import openpyxl
import asyncio
from bs4 import BeautifulSoup
from zipfile import ZipFile
import os


async def fetch(url):
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            return await response.text()

# Getting Paginations Links
async def get_pageinations_links(soup):
    # Getting Pagination for number of pages should iterate
    pages_iter = soup.find("ul", class_ = 'pagination')
    try:
        pages_iter_li = pages_iter.find_all("li")
        # print("Number of pages:", len(pages_iter_li))

        # Extract href attributes from <a> tags inside each <li>
        href_pages = [li.find('a')['href'] for li in pages_iter_li if li.find('a')]
        # print("Pagination hrefs:", href_pages)
        return href_pages
    except Exception as e:
        print("No pagination found", e)
        return []
    
    
# Finding all the heading of tables
async def get_table_headers(soup):
    headers_text = soup.find_all('th')
    main_table_headers = [title.text.strip() for title in headers_text]
    # print(main_table_headers) # ['Team Name', 'Year', 'Wins', 'Losses', 'OT Losses', 'Win %', 'Goals For (GF)', 'Goals Against (GA)', '+ / -']
    return main_table_headers

# Function to fetch HTML content from the given URL
async def fetch_html(url):
    # response = requests.get(url)
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as resp:
            # print(resp.status)
            # print(await resp.text())
            if resp.status == 200:
                return await resp.text()
            else:
                print(f"Failed to fetch URL: {url}, Status Code: {resp.status}")
                return None
    
async def save_html_pages(soups, page_count, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    filenaem = f"{page_count}.html" 
    file_path = os.path.join(output_dir, filenaem)
    with open(file_path, "w", encoding= 'utf-8') as html:
        html.write(str(soups.prettify()))

async def save_data_sheet(soups, data):
    cell_datas = soups.find_all('tr', class_='team')
    for row_data in cell_datas:
        try:
            name = row_data.find("td", class_="name").text.strip()
            year = row_data.find("td", class_="year").text.strip()
            wins = row_data.find("td", class_="wins").text.strip()
            losses = row_data.find("td", class_="losses").text.strip()
            ot_losses = row_data.find("td", class_="ot-losses").text.strip()
            pct = row_data.find("td", class_="pct").text.strip()
            gf = row_data.find("td", class_="gf").text.strip()
            ga = row_data.find("td", class_="ga").text.strip()
            diff = row_data.find("td", class_="diff").text.strip()
            data.append((name, year, wins, losses, ot_losses, pct, gf, ga, diff))
        except AttributeError as e:
            print("Error processing row:", e)


async def zip_html_pages(zip_dir, output_dir):
    with ZipFile(zip_dir, 'w') as ziper:
        for root, _, files in os.walk(output_dir):
            for file in files:
                ziper.write(os.path.join(root, file), arcname=file)


async def write_to_excel(data_dicts, main_table_headers, file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    # Write headers
    sheet.append(main_table_headers)
    # Write rows in sheet
    for data_dict in data_dicts:  # Assuming data_dicts is a list of dictionaries
        sheet.append([data_dict.get(header, "") for header in main_table_headers])
    # Save the workbook
    workbook.save("output.xlsx")


async def calculate_winner_loser(file_name, output_sheet_name):
    worksheet_name =  r"output.xlsx"
    rb = openpyxl.load_workbook(worksheet_name)
    sheet = rb.active
    ws = rb.create_sheet('Winner and Loser per Year')
    # Create headers for the new sheet
    ws.append(['Year', 'Winner', 'Winner Num. of Wins', 'Loser', 'Loser Num. of Wins'])
    temp = []
    for row in sheet:
        year = row[1].value
        name = row[0].value
        wins = row[2].value
        # print(year, name, wins)
        temp.append((year,name,wins))

    #  Getting unique years from temp list
    years_temp = []
    x = [years_temp.append(temp[x][0]) for x in range(len(temp)) if temp[x][0] not in years_temp]
    # print("Unique Years: ", years_temp)
    for year in years_temp[1:]:
        max_ls = []
        for data in temp:
            if year == data[0]:
                max_ls.append(data)
                
        max_value = max(max_ls, key=lambda x: int(x[2]))
        min_value = min(max_ls, key=lambda x: int(x[2]))
        
        year_max_min = max_value[0]
        team_max = max_value[1]
        team_min = min_value[1]
        score_max = max_value[2]
        score_min = min_value[2]
        # print(f"Year: {year_max_min} | Max_Team: {team_max} | MAXScore: {score_max} | Min_team: {team_min} | MinScore: {score_min}")
        ws.append([year_max_min, team_max, score_max, team_min, score_min])
    
    rb.save(worksheet_name)
    print(f"Sheet '{output_sheet_name}' updated successfully!")


#########  Main Fucntion #########
async def main():
    # Main URL to fetch
    url = "https://www.scrapethissite.com/pages/forms/"
    response = await fetch(url)
    # print("AIOHTTP: ", response)
    soup = BeautifulSoup(response, "html.parser")
    main_table_headers = await get_table_headers(soup)
    pagination_links = await get_pageinations_links(soup)
    
    zip_dir = "Collection_html.zip"
    output_dir = "HTML PAGES"
    file_name = "output.xlsx"
    page_count = 1
    data = []
    
    
    for page in pagination_links:
        page_url = f"https://www.scrapethissite.com/{page}"
        soups = BeautifulSoup(await fetch_html(page_url), 'html.parser')
        await save_html_pages(soups, page_count, output_dir)
        await save_data_sheet(soups, data)
        page_count += 1

    
    await zip_html_pages(zip_dir, output_dir)
    
    # Create dictionary list for writing to Excel
    data_dicts = [dict(zip(main_table_headers, row)) for row in data]
    
    await write_to_excel(data_dicts, main_table_headers, file_name)
    
    # Calculate and append winners/losers per year
    await calculate_winner_loser(file_name, "Winner and Loser per Year")
    
if __name__ == "__main__":
    asyncio.run(main())
